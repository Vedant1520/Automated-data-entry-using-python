import openpyxl
import os.path
import sys

input_file = ""
output_file=""
name=""



def pageCode(column_names, values):
      output = """<!doctype html>
	<html>
	<head>
	<meta charset="utf-8">
	<title>{first_column_name} - {first_column_value}</title>
	</head>
	<body>
	<table width="500px" border="1">
	<tbody>""".format(first_column_name = column_names[0], first_column_value = values[0])

      for i in range(0,len(column_names)):
        output += '<tr><th scope="row">{column_name}</th><td>{column_value}</td></tr>'.format(column_name = column_names[i], column_value = values[i])

      output += """</tbody>
      </table>
      </body>
      </html>"""
      
      return output

def run(): 
      if(os.path.isfile(input_file) == False):
            print("The Excel File Does Not Exist ")
            sys.exit()
      else:
            read = openpyxl.load_workbook(input_file)
            sheet = read.active
            read.close()

      if(os.path.isdir(output_file) is False):
            os.mkdir(output_file)
      else:
            if(os.listdir(output_file)!=[]):
                  print("The Given Output Folder Is Not Empty!!")
                  print("Clear the folder and try again !!") 
                  sys.exit()

      column_names = []
      for i in range(1, sheet.max_column+1):
        column_names.append(sheet.cell(row=1, column=i).value)

      for i in range(2, sheet.max_row+1):
        values = []
        for j in range(1, sheet.max_column+1):
            values.append(sheet.cell(row=i, column=j).value)
        createFile(column_names, values)



def createFile(column_names, values):
    
      filename =output_file+name+"- {Name}.html".format(Name=values[0]) 
      f = open(filename,"w")
      f.write(pageCode(column_names, values)) 
      f.close()   

